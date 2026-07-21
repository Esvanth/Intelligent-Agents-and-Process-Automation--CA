"""RPA component: log applicant metadata to an Excel tracker."""
from pathlib import Path
from openpyxl import Workbook, load_workbook

TRACKER_PATH = Path("data/tracker.xlsx")
HEADERS = ["Received At", "Candidate Name", "Candidate Email", "Subject",
           "CV Path", "AI Score", "Recommendation", "Status"]


def _init_tracker():
    if TRACKER_PATH.exists(): return
    wb = Workbook()
    ws = wb.active
    ws.title = "Applicants"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    widths = [22, 18, 28, 30, 40, 10, 22, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    wb.save(TRACKER_PATH)


def log_applicant(metadata: dict) -> int:
    _init_tracker()
    wb = load_workbook(TRACKER_PATH)
    ws = wb["Applicants"]
    row = [
        metadata.get("received_at", ""),
        metadata.get("candidate_name", ""),
        metadata.get("candidate_email", ""),
        metadata.get("subject", ""),
        metadata.get("cv_path", ""),
        "", "", "Received",
    ]
    ws.append(row)
    wb.save(TRACKER_PATH)
    return ws.max_row


def update_score(candidate_email: str, score: float, recommendation: str):
    wb = load_workbook(TRACKER_PATH)
    ws = wb["Applicants"]
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=3).value == candidate_email:
            ws.cell(row=r, column=6).value = score
            ws.cell(row=r, column=7).value = recommendation
            ws.cell(row=r, column=8).value = "Scored"
            break
    wb.save(TRACKER_PATH)


def mark_status(candidate_email: str, status: str):
    wb = load_workbook(TRACKER_PATH)
    ws = wb["Applicants"]
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=3).value == candidate_email:
            ws.cell(row=r, column=8).value = status
            break
    wb.save(TRACKER_PATH)
